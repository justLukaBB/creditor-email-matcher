"""
XLSX Document Extractor

Extracts data from XLSX spreadsheets using openpyxl in memory-efficient read_only mode.
Searches for amount keywords and extracts adjacent numeric values.

Phase 3 Scope:
- Memory-efficient extraction with read_only=True and data_only=True
- Keyword-based amount detection (Gesamtforderung, Betrag, etc.)
- German number format parsing
- No API calls - pure openpyxl extraction
"""

import os
import re
import logging
from typing import List, Dict, Any

from app.models.extraction_result import (
    SourceExtractionResult,
    ExtractedAmount,
    ExtractedEntity,
)
from app.services.extraction.german_preprocessor import GermanTextPreprocessor
from app.services.extraction.german_parser import parse_german_amount
from app.services.extraction.german_validator import GermanValidator

logger = logging.getLogger(__name__)


class XLSXExtractor:
    """
    Extracts structured data from XLSX spreadsheets.

    Uses openpyxl with read_only=True for memory efficiency (constant memory).
    Searches for amount keywords and extracts adjacent cell values.
    """

    def __init__(self):
        self.preprocessor = GermanTextPreprocessor()
        self.validator = GermanValidator()
        # Keywords indicating amount values (lowercase for comparison)
        self.amount_keywords = [
            'gesamtforderung', 'gesamt', 'forderung', 'betrag', 'summe',
            'total', 'amount', 'schuld', 'forderungshöhe', 'forderungshoehe',
            'hauptforderung', 'nebenforderung', 'zinsen', 'kosten'
        ]

    def extract(self, xlsx_path: str) -> SourceExtractionResult:
        """
        Extract claim amounts from XLSX spreadsheet.

        Args:
            xlsx_path: Path to the XLSX file

        Returns:
            SourceExtractionResult with extracted amounts
        """
        from openpyxl import load_workbook

        result = SourceExtractionResult(
            source_type="xlsx",
            source_name=os.path.basename(xlsx_path),
            extraction_method="openpyxl",
            tokens_used=0
        )

        try:
            # Memory-efficient mode: read_only=True, data_only=True
            # read_only=True: streams data instead of loading entire file
            # data_only=True: reads cached values instead of formulas
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)

            found_amounts = []
            sheet_count = 0
            rows_processed = 0

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_count += 1

                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    rows_processed += 1

                    # Skip empty rows
                    if not any(cell is not None for cell in row):
                        continue

                    # Convert row to list for indexing
                    row_list = list(row)

                    # Look for keyword + value pattern
                    for col_idx, cell in enumerate(row_list):
                        if cell is None:
                            continue

                        # NEW: Apply preprocessing to cell text
                        cell_str = str(cell)
                        cell_str = self.preprocessor.preprocess(cell_str)
                        cell_str_lower = cell_str.lower()

                        # Check if this cell contains a keyword
                        if any(kw in cell_str_lower for kw in self.amount_keywords):
                            # Look at adjacent cells for numeric value
                            for adjacent_idx in [col_idx + 1, col_idx - 1]:
                                if 0 <= adjacent_idx < len(row_list):
                                    adjacent_cell = row_list[adjacent_idx]
                                    if adjacent_cell is not None:
                                        amount_info = self._parse_amount(adjacent_cell)
                                        if amount_info:
                                            found_amounts.append({
                                                'value': amount_info['value'],
                                                'raw': f"{cell}: {adjacent_cell}",
                                                'sheet': sheet_name,
                                                'row': row_idx,
                                                'confidence': amount_info['confidence']
                                            })

            # Always close workbook when done
            wb.close()

            logger.info(
                f"XLSXExtractor: processed {sheet_count} sheets, "
                f"{rows_processed} rows, found {len(found_amounts)} amounts "
                f"in {os.path.basename(xlsx_path)}"
            )

            # Take highest amount (USER DECISION: highest wins)
            if found_amounts:
                best = max(found_amounts, key=lambda x: x['value'])
                result.gesamtforderung = ExtractedAmount(
                    value=best['value'],
                    currency="EUR",
                    raw_text=best['raw'],
                    source="xlsx",
                    confidence=best['confidence']
                )
                logger.info(
                    f"XLSXExtractor: selected {best['value']} EUR "
                    f"(confidence: {best['confidence']}) from sheet '{best['sheet']}'"
                )

            return result

        except Exception as e:
            result.error = f"xlsx_extraction_error: {str(e)}"
            logger.error(f"XLSXExtractor: failed to extract from {xlsx_path}: {e}")
            return result

    def _parse_amount(self, cell_value: Any) -> Dict[str, Any] | None:
        """
        Parse a cell value as a monetary amount.

        Handles:
        - Native numeric types (int, float)
        - German format strings ("1.234,56")
        - Currency prefixes/suffixes ("€", "EUR")

        Returns dict with 'value' and 'confidence' or None if not parseable.
        """
        try:
            # Direct numeric value
            if isinstance(cell_value, (int, float)):
                if cell_value > 0:
                    return {
                        'value': float(cell_value),
                        'confidence': 'HIGH'  # Native numbers are reliable
                    }
                return None

            # String parsing
            amount_str = str(cell_value).strip()

            # NEW: Use babel-based parser instead of manual replacement
            try:
                amount_value = parse_german_amount(amount_str)
                if amount_value > 0:
                    # Check for German format (comma as decimal separator)
                    has_german_format = ',' in amount_str
                    return {
                        'value': amount_value,
                        'confidence': 'HIGH' if has_german_format else 'MEDIUM'
                    }
            except ValueError:
                pass  # Fall through to return None

        except (ValueError, TypeError):
            pass

        return None
